"""Text extraction utilities for various sources.

Supports extraction from:
- Files (PDF, DOCX, ODT, HTML, text files, CSV, JSON, XML, YAML, TOML, INI, XLSX, ODS)
- Archives (ZIP, TAR, TAR.GZ, TAR.BZ2)
- URLs (web pages, documents)
- Python objects (classes, functions, modules, instances)
- File system (directory listings)
- Email files (EML)
"""

from typing import Any, Optional, Dict, List
import os
import json
import inspect
import re
from io import BytesIO, StringIO, TextIOWrapper
import csv
import xml.etree.ElementTree as ET
import configparser
import zipfile
import tarfile
import email
from email import policy
from pathlib import Path

# For gitignore handling
try:
    import pathspec
    HAS_PATHSPEC = True
except ImportError:
    HAS_PATHSPEC = False

# Try to import optional stdlib modules
try:
    import tomllib  # Python 3.11+
except ImportError:
    try:
        import tomli as tomllib  # Fallback for older Python
    except ImportError:
        tomllib = None

import requests
from bs4 import BeautifulSoup
import PyPDF2
import docx
from odf import opendocument, text, teletype
import yaml

# Optional imports for extended functionality
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import trafilatura
    HAS_TRAFILATURA = True
except ImportError:
    HAS_TRAFILATURA = False

from .get_webdriver import get_webdriver


# ============================================================================
# Text Processing Utilities
# ============================================================================

def strip_newlines(string: str) -> str:
    """Remove leading/trailing newlines and collapse multiple newlines.

    Args:
        string: Input text to clean

    Returns:
        Cleaned text with normalized newlines
    """
    string = string.strip('\n')
    return re.sub(r'\n{2,}', '\n', string)


# ============================================================================
# Document Processing
# ============================================================================

def _extract_pdf_text(file_obj) -> str:
    """Extract text from PDF file object.

    Args:
        file_obj: File-like object containing PDF data

    Returns:
        Extracted text from all pages
    """
    reader = PyPDF2.PdfReader(file_obj)
    return "".join(page.extract_text() for page in reader.pages)


def _extract_docx_text(file_obj) -> str:
    """Extract text from DOCX file object.

    Args:
        file_obj: File-like object or path to DOCX file

    Returns:
        Extracted text from all paragraphs
    """
    doc = docx.Document(file_obj)
    return "\n".join(para.text for para in doc.paragraphs)


def _extract_odt_text(file_obj) -> str:
    """Extract text from ODT file object.

    Args:
        file_obj: File-like object or path to ODT file

    Returns:
        Extracted text from all paragraphs
    """
    doc = opendocument.load(file_obj)
    paragraphs = doc.getElementsByType(text.P)
    return '\n'.join(teletype.extractText(para) for para in paragraphs)


def _extract_html_text(file_obj) -> str:
    """Extract text from HTML file object.

    Args:
        file_obj: File-like object containing HTML

    Returns:
        Extracted text content
    """
    soup = BeautifulSoup(file_obj, 'html.parser')
    return soup.get_text()


def _extract_csv_text(file_obj) -> str:
    """Extract text from CSV file object.

    Args:
        file_obj: File-like object containing CSV data

    Returns:
        Formatted CSV content as text
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8')

    reader = csv.reader(StringIO(content))
    lines = []
    for row in reader:
        lines.append(' | '.join(row))
    return '\n'.join(lines)


def _extract_json_text(file_obj) -> str:
    """Extract text from JSON file object.

    Args:
        file_obj: File-like object containing JSON data

    Returns:
        Pretty-printed JSON content
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8')

    data = json.loads(content)
    return json.dumps(data, indent=2, ensure_ascii=False)


def _extract_xml_text(file_obj) -> str:
    """Extract text from XML file object.

    Args:
        file_obj: File-like object containing XML data

    Returns:
        XML content as formatted text
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8')

    try:
        tree = ET.ElementTree(ET.fromstring(content))
        root = tree.getroot()

        def extract_element(elem, level=0):
            """Recursively extract text from XML elements."""
            indent = '  ' * level
            text_parts = []

            # Add element tag
            text_parts.append(f"{indent}<{elem.tag}>")

            # Add element text if present
            if elem.text and elem.text.strip():
                text_parts.append(f"{indent}  {elem.text.strip()}")

            # Recurse into children
            for child in elem:
                text_parts.append(extract_element(child, level + 1))

            return '\n'.join(text_parts)

        return extract_element(root)
    except ET.ParseError:
        # If parsing fails, return raw content
        return content


def _extract_yaml_text(file_obj) -> str:
    """Extract text from YAML file object.

    Args:
        file_obj: File-like object containing YAML data

    Returns:
        Pretty-printed YAML content
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8')

    data = yaml.safe_load(content)
    return yaml.dump(data, default_flow_style=False, allow_unicode=True)


def _extract_toml_text(file_obj) -> str:
    """Extract text from TOML file object.

    Args:
        file_obj: File-like object containing TOML data

    Returns:
        TOML content as text (raw if tomllib not available)
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8')

    if tomllib is None:
        # If tomllib not available, return raw content
        return content

    data = tomllib.loads(content)
    # Since there's no standard toml dump in stdlib, return pretty JSON representation
    return json.dumps(data, indent=2, ensure_ascii=False)


def _extract_ini_text(file_obj) -> str:
    """Extract text from INI/CFG file object.

    Args:
        file_obj: File-like object containing INI data

    Returns:
        Formatted INI content as text
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8')

    config = configparser.ConfigParser()
    config.read_string(content)

    lines = []
    for section in config.sections():
        lines.append(f"[{section}]")
        for key, value in config.items(section):
            lines.append(f"{key} = {value}")
        lines.append("")

    return '\n'.join(lines)


def _extract_xlsx_text(file_obj) -> str:
    """Extract text from XLSX file object.

    Args:
        file_obj: File-like object or path to XLSX file

    Returns:
        Formatted spreadsheet content as text
    """
    if not HAS_OPENPYXL:
        return "Error: openpyxl not installed. Install with: pip install openpyxl"

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    lines = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        lines.append(f"=== Sheet: {sheet_name} ===\n")

        for row in sheet.iter_rows(values_only=True):
            row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
            if row_text.strip():
                lines.append(row_text)

        lines.append("")

    return '\n'.join(lines)


def _extract_ods_text(file_obj) -> str:
    """Extract text from ODS spreadsheet file object.

    Args:
        file_obj: File-like object or path to ODS file

    Returns:
        Formatted spreadsheet content as text
    """
    doc = opendocument.load(file_obj)
    from odf.table import Table, TableRow, TableCell

    lines = []
    tables = doc.getElementsByType(Table)

    for table in tables:
        table_name = table.getAttribute('name')
        if table_name:
            lines.append(f"=== Sheet: {table_name} ===\n")

        rows = table.getElementsByType(TableRow)
        for row in rows:
            cells = row.getElementsByType(TableCell)
            row_data = []
            for cell in cells:
                cell_text = teletype.extractText(cell)
                row_data.append(cell_text)

            row_text = ' | '.join(row_data)
            if row_text.strip():
                lines.append(row_text)

        lines.append("")

    return '\n'.join(lines)


def _extract_eml_text(file_obj) -> str:
    """Extract text from email file object.

    Args:
        file_obj: File-like object containing email data

    Returns:
        Formatted email content as text
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        msg = email.message_from_bytes(content, policy=policy.default)
    else:
        msg = email.message_from_string(content, policy=policy.default)

    lines = []

    # Extract headers
    lines.append("=== Email Headers ===")
    lines.append(f"From: {msg.get('From', 'Unknown')}")
    lines.append(f"To: {msg.get('To', 'Unknown')}")
    lines.append(f"Subject: {msg.get('Subject', 'No Subject')}")
    lines.append(f"Date: {msg.get('Date', 'Unknown')}")
    lines.append("")

    # Extract body
    lines.append("=== Email Body ===")
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            if content_type == 'text/plain':
                lines.append(part.get_content())
    else:
        lines.append(msg.get_content())

    return '\n'.join(lines)


# ============================================================================
# Archive Handlers
# ============================================================================

def _extract_zip_text(file_obj) -> str:
    """Extract file listing from ZIP archive.

    Args:
        file_obj: File-like object or path to ZIP file

    Returns:
        List of files in the archive with sizes
    """
    if isinstance(file_obj, str):
        zf = zipfile.ZipFile(file_obj, 'r')
    else:
        zf = zipfile.ZipFile(file_obj)

    lines = ["=== ZIP Archive Contents ===\n"]

    for info in zf.infolist():
        size_str = f"{info.file_size:,} bytes"
        lines.append(f"{info.filename:<50} {size_str:>15}")

    lines.append(f"\nTotal files: {len(zf.filelist)}")

    return '\n'.join(lines)


def _extract_tar_text(file_obj, mode='r') -> str:
    """Extract file listing from TAR archive.

    Args:
        file_obj: File-like object or path to TAR file
        mode: Open mode ('r', 'r:gz', 'r:bz2', etc.)

    Returns:
        List of files in the archive with sizes
    """
    if isinstance(file_obj, str):
        tf = tarfile.open(file_obj, mode)
    else:
        tf = tarfile.open(fileobj=file_obj, mode=mode)

    lines = ["=== TAR Archive Contents ===\n"]

    for member in tf.getmembers():
        size_str = f"{member.size:,} bytes"
        type_char = 'd' if member.isdir() else 'f'
        lines.append(f"{type_char} {member.name:<50} {size_str:>15}")

    lines.append(f"\nTotal entries: {len(tf.getmembers())}")

    return '\n'.join(lines)


# ============================================================================
# File System Handlers
# ============================================================================

def _load_gitignore_patterns(directory: str) -> Optional['pathspec.PathSpec']:
    """Load gitignore patterns from default and local .gitignore files.

    Args:
        directory: Directory to scan for local .gitignore

    Returns:
        PathSpec object with combined patterns, or None if pathspec not available
    """
    if not HAS_PATHSPEC:
        return None

    patterns = []

    # Load default gitignore patterns bundled with get_text
    default_gitignore = Path(__file__).parent / 'default_gitignore'
    if default_gitignore.exists():
        with open(default_gitignore, 'r', encoding='utf-8') as f:
            patterns.extend(f.readlines())

    # Load local .gitignore if present
    local_gitignore = Path(directory) / '.gitignore'
    if local_gitignore.exists():
        with open(local_gitignore, 'r', encoding='utf-8') as f:
            patterns.extend(f.readlines())

    # Create PathSpec from patterns
    if patterns:
        return pathspec.PathSpec.from_lines('gitwildmatch', patterns)

    return None


def handle_directory(path: str, max_depth: int = 3) -> str:
    """Generate a tree-like representation of directory structure.

    Automatically applies sensible exclusions from:
    - Default gitignore patterns (bundled with get_text)
    - Local .gitignore file (if present)

    This filters out common junk directories like .git, __pycache__, node_modules, etc.
    Recursion is limited to max_depth levels to avoid overwhelming output.

    Args:
        path: Path to directory
        max_depth: Maximum recursion depth (default: 3)

    Returns:
        Tree structure as text, or error message
    """
    if not os.path.exists(path):
        return f'The path {path} does not exist.'

    if not os.path.isdir(path):
        return f'The path {path} is not a valid directory.'

    # Load gitignore patterns
    spec = _load_gitignore_patterns(path)

    # Track directories that weren't fully explored
    unexplored_dirs = []

    def should_ignore(item_path: str, base_path: str) -> bool:
        """Check if path should be ignored based on gitignore patterns."""
        if spec is None:
            return False

        # Get relative path from base
        try:
            rel_path = os.path.relpath(item_path, base_path)
            # pathspec expects forward slashes
            rel_path = rel_path.replace(os.sep, '/')
            # Check if it's a directory
            if os.path.isdir(item_path):
                rel_path += '/'
            return spec.match_file(rel_path)
        except ValueError:
            # Handles case where paths are on different drives on Windows
            return False

    def has_contents(folder: str) -> bool:
        """Check if a directory has non-ignored contents."""
        try:
            contents = os.listdir(folder)
            for item in contents:
                item_path = os.path.join(folder, item)
                if not should_ignore(item_path, path):
                    return True
            return False
        except PermissionError:
            return False

    def recurse_folder(folder: str, prefix: str = '', base_path: str = None, depth: int = 0) -> str:
        """Recursively build directory tree with gitignore filtering and depth limit."""
        if base_path is None:
            base_path = folder

        try:
            contents = os.listdir(folder)
        except PermissionError:
            return prefix + '├── [Permission Denied]\n'

        # Filter out ignored items
        filtered_contents = []
        for item in contents:
            item_path = os.path.join(folder, item)
            if not should_ignore(item_path, base_path):
                filtered_contents.append(item)

        output = ''
        for i, item in enumerate(filtered_contents):
            item_path = os.path.join(folder, item)
            is_dir = os.path.isdir(item_path)

            # Check if we've reached max depth for this directory
            depth_limit_reached = is_dir and depth >= max_depth

            # Check if directory has unexplored contents
            has_more = is_dir and depth_limit_reached and has_contents(item_path)

            # Format the item name with appropriate indicators
            if has_more:
                item_display = f"{item}/ [...explore further]"
                unexplored_dirs.append(item_path)
            elif is_dir and depth < max_depth:
                # Check if empty after filtering
                if not has_contents(item_path):
                    item_display = f"{item}/ [empty]"
                else:
                    item_display = f"{item}/"
            elif is_dir:
                item_display = f"{item}/"
            else:
                item_display = item

            output += prefix + '├── ' + item_display + '\n'

            # Recurse into subdirectories if under depth limit
            if is_dir and depth < max_depth:
                new_prefix = prefix + ('    ' if i == len(filtered_contents) - 1 else '│   ')
                output += recurse_folder(item_path, new_prefix, base_path, depth + 1)

        return output

    result = path + '\n' + recurse_folder(path, depth=0)

    # Add notes about filtering and depth limit
    notes = []

    if spec is not None:
        notes.append("Common junk directories (.git, __pycache__, node_modules, etc.) are automatically filtered")
    elif HAS_PATHSPEC:
        notes.append("No gitignore patterns found, showing all files")
    else:
        notes.append("Warning: pathspec not installed. Install with 'pip install pathspec' for gitignore filtering")

    notes.append(f"Recursion limited to {max_depth} levels deep")

    if unexplored_dirs:
        notes.append(f"Directories with unexplored contents ({len(unexplored_dirs)} total):")
        for unexplored_dir in unexplored_dirs[:10]:  # Show max 10 examples
            notes.append(f"  - {unexplored_dir}")
        if len(unexplored_dirs) > 10:
            notes.append(f"  ... and {len(unexplored_dirs) - 10} more")
        notes.append("Tip: Call get_text() on specific subdirectories to explore them further")

    result += '\n\n[' + '\n '.join(notes) + ']'

    return result


def handle_file(source: str) -> str:
    """Extract text from a file based on its extension.

    Supported formats:
    - Documents: PDF, DOCX, ODT, HTML, TXT
    - Data: CSV, JSON, XML, YAML, TOML, INI, CFG
    - Spreadsheets: XLSX, ODS
    - Archives: ZIP, TAR, TAR.GZ, TAR.BZ2, TGZ, TBZ2
    - Email: EML

    Args:
        source: Path to file

    Returns:
        Extracted text content, or error message
    """
    ext = os.path.splitext(source)[1].lower()

    try:
        # Document formats
        if ext == '.pdf':
            with open(source, 'rb') as file:
                return _extract_pdf_text(file)

        elif ext == '.docx':
            return _extract_docx_text(source)

        elif ext == '.odt':
            return _extract_odt_text(source)

        elif ext in ('.html', '.htm'):
            with open(source, 'r', encoding='utf-8') as f:
                return _extract_html_text(f)

        # Data formats
        elif ext == '.csv':
            with open(source, 'rb') as f:
                return _extract_csv_text(f)

        elif ext == '.json':
            with open(source, 'rb') as f:
                return _extract_json_text(f)

        elif ext == '.xml':
            with open(source, 'rb') as f:
                return _extract_xml_text(f)

        elif ext in ('.yaml', '.yml'):
            with open(source, 'rb') as f:
                return _extract_yaml_text(f)

        elif ext == '.toml':
            with open(source, 'rb') as f:
                return _extract_toml_text(f)

        elif ext in ('.ini', '.cfg', '.conf'):
            with open(source, 'rb') as f:
                return _extract_ini_text(f)

        # Spreadsheet formats
        elif ext == '.xlsx':
            return _extract_xlsx_text(source)

        elif ext == '.ods':
            return _extract_ods_text(source)

        # Archive formats
        elif ext == '.zip':
            return _extract_zip_text(source)

        elif ext in ('.tar', '.tar.gz', '.tgz'):
            mode = 'r:gz' if ext in ('.tar.gz', '.tgz') else 'r'
            with open(source, 'rb') as f:
                return _extract_tar_text(f, mode=mode)

        elif ext in ('.tar.bz2', '.tbz2'):
            with open(source, 'rb') as f:
                return _extract_tar_text(f, mode='r:bz2')

        # Email format
        elif ext == '.eml':
            with open(source, 'rb') as f:
                return _extract_eml_text(f)

        else:
            # Plain text file
            with open(source, 'r', encoding='utf-8') as f:
                return f.read()

    except FileNotFoundError:
        return f"Unable to process file. File not found: {source}"
    except PermissionError:
        return f"Unable to process file. Permission denied: {source}"
    except Exception as e:
        return f"Error while attempting to read file: {e}"


# ============================================================================
# Web Handlers
# ============================================================================

def _extract_with_beautifulsoup(html_content: str) -> str:
    """Extract text from HTML using BeautifulSoup with smart filtering.

    Removes common non-content elements like scripts, styles, navigation, etc.

    Args:
        html_content: HTML content as string

    Returns:
        Extracted text content
    """
    soup = BeautifulSoup(html_content, 'html.parser')

    # Remove non-content elements
    for element in soup(['script', 'style', 'nav', 'header', 'footer',
                        'aside', 'iframe', 'noscript']):
        element.decompose()

    # Try to find main content area first
    main_content = soup.find('main') or soup.find('article') or soup.find('body')

    if main_content:
        text = main_content.get_text(separator='\n', strip=True)
    else:
        text = soup.get_text(separator='\n', strip=True)

    return strip_newlines(text)


def _is_content_sufficient(text: str, min_length: int = 100) -> bool:
    """Check if extracted content seems sufficient (not empty or too short).

    Args:
        text: Extracted text content
        min_length: Minimum acceptable length in characters

    Returns:
        True if content seems sufficient, False otherwise
    """
    if not text or len(text.strip()) < min_length:
        return False

    # Check if content is not just boilerplate
    words = text.split()
    if len(words) < 20:  # Less than 20 words is suspicious
        return False

    return True


def extract_webpage_content(url: str, force_selenium: bool = False) -> str:
    """Extract text content from a web page using a hybrid approach.

    Strategy:
    1. If force_selenium=False:
       - Try trafilatura first (if available) - fast and smart
       - If trafilatura unavailable or content insufficient, try BeautifulSoup
       - If still insufficient, fallback to Selenium (handles JavaScript)
    2. If force_selenium=True:
       - Use Selenium directly (for known dynamic pages)

    Args:
        url: URL to fetch
        force_selenium: If True, skip lightweight methods and use Selenium directly

    Returns:
        Extracted text content from the page
    """
    # Force Selenium path
    if force_selenium:
        driver = None
        try:
            driver = get_webdriver()
            driver.get(url)
            page_source = driver.page_source
            return _extract_with_beautifulsoup(page_source)
        finally:
            if driver:
                driver.quit()

    # Try lightweight methods first
    try:
        # Fetch page content
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        html_content = response.text

        # Method 1: Try trafilatura if available (best for articles/blogs)
        if HAS_TRAFILATURA:
            extracted = trafilatura.extract(html_content, include_comments=False,
                                           include_tables=True)
            if extracted and _is_content_sufficient(extracted):
                return strip_newlines(extracted)

        # Method 2: Try BeautifulSoup with smart filtering
        text = _extract_with_beautifulsoup(html_content)
        if _is_content_sufficient(text):
            return text

        # Method 3: Content seems insufficient, might need JavaScript
        # Fallback to Selenium
        driver = None
        try:
            driver = get_webdriver()
            driver.get(url)
            page_source = driver.page_source
            return _extract_with_beautifulsoup(page_source)
        finally:
            if driver:
                driver.quit()

    except requests.exceptions.RequestException:
        # If requests fails, try Selenium as last resort
        driver = None
        try:
            driver = get_webdriver()
            driver.get(url)
            page_source = driver.page_source
            return _extract_with_beautifulsoup(page_source)
        finally:
            if driver:
                driver.quit()


def handle_url(source: str) -> str:
    """Extract text from a URL (document or web page).

    Automatically handles various document types served over HTTP based on
    content-type or URL extension, falling back to web scraping for HTML pages.

    Supported formats:
    - Documents: PDF, DOCX, ODT
    - Data: CSV, JSON, XML, YAML
    - Spreadsheets: XLSX, ODS
    - Archives: ZIP
    - Web pages: HTML (via Selenium)

    Args:
        source: URL to fetch

    Returns:
        Extracted text content, or error message
    """
    try:
        response = requests.get(source, timeout=30)
        response.raise_for_status()

        content_type = response.headers.get('content-type', '').lower()

        # Handle PDF documents
        if 'application/pdf' in content_type or source.endswith('.pdf'):
            with BytesIO(response.content) as file:
                return _extract_pdf_text(file)

        # Handle DOCX documents
        elif 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' in content_type \
             or source.endswith('.docx'):
            with BytesIO(response.content) as file:
                return _extract_docx_text(file)

        # Handle ODT documents
        elif 'application/vnd.oasis.opendocument.text' in content_type \
             or source.endswith('.odt'):
            with BytesIO(response.content) as file:
                return _extract_odt_text(file)

        # Handle XLSX spreadsheets
        elif 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type \
             or source.endswith('.xlsx'):
            with BytesIO(response.content) as file:
                return _extract_xlsx_text(file)

        # Handle ODS spreadsheets
        elif 'application/vnd.oasis.opendocument.spreadsheet' in content_type \
             or source.endswith('.ods'):
            with BytesIO(response.content) as file:
                return _extract_ods_text(file)

        # Handle CSV
        elif 'text/csv' in content_type or source.endswith('.csv'):
            with BytesIO(response.content) as file:
                return _extract_csv_text(file)

        # Handle JSON
        elif 'application/json' in content_type or source.endswith('.json'):
            with BytesIO(response.content) as file:
                return _extract_json_text(file)

        # Handle XML
        elif 'application/xml' in content_type or 'text/xml' in content_type \
             or source.endswith('.xml'):
            with BytesIO(response.content) as file:
                return _extract_xml_text(file)

        # Handle YAML
        elif 'application/x-yaml' in content_type or source.endswith(('.yaml', '.yml')):
            with BytesIO(response.content) as file:
                return _extract_yaml_text(file)

        # Handle ZIP archives
        elif 'application/zip' in content_type or source.endswith('.zip'):
            with BytesIO(response.content) as file:
                return _extract_zip_text(file)

        # Default to web scraping
        else:
            return extract_webpage_content(source)

    except requests.exceptions.Timeout:
        return f"Unable to process URL. Request timed out: {source}"
    except requests.exceptions.RequestException as e:
        return f"Unable to process URL. Connection error: {e}"


# ============================================================================
# Python Object Inspection
# ============================================================================

def handle_class(source: type) -> str:
    """Generate detailed JSON representation of a class.

    Args:
        source: Class object to inspect

    Returns:
        JSON string with class metadata
    """
    class_info = {
        "type": "class",
        "name": source.__name__,
        "doc": inspect.getdoc(source) or "No documentation available.",
        "base_classes": [base.__name__ for base in inspect.getmro(source)[1:]],
        "methods": {},
        "class_methods": {},
        "static_methods": {},
        "properties": {},
        "attributes": {},
        "string_repr": repr(source)
    }

    # Process methods, class methods, static methods, and properties
    for name, member in inspect.getmembers(source):
        if inspect.isfunction(member):
            class_info["methods"][name] = {
                "doc": inspect.getdoc(member) or "No documentation available.",
                "signature": str(inspect.signature(member))
            }
        elif isinstance(member, classmethod):
            class_info["class_methods"][name] = {
                "doc": inspect.getdoc(member) or "No documentation available.",
                "signature": str(inspect.signature(member.__func__))
            }
        elif isinstance(member, staticmethod):
            class_info["static_methods"][name] = {
                "doc": inspect.getdoc(member) or "No documentation available."
            }
        elif isinstance(member, property):
            class_info["properties"][name] = {
                "doc": inspect.getdoc(member) or "No documentation available."
            }

    # Process attributes
    for name in dir(source):
        if not name.startswith('__') and not inspect.isroutine(getattr(source, name)):
            try:
                attribute = inspect.getattr_static(source, name)
                serialized = json.dumps(attribute)
            except:
                serialized = repr(attribute)

            class_info["attributes"][name] = {
                "value": serialized,
                "type": type(attribute).__name__
            }

    return json.dumps(class_info, indent=4)


def handle_function(source: Any) -> str:
    """Generate detailed JSON representation of a function or callable.

    Args:
        source: Function, method, or callable object to inspect

    Returns:
        JSON string with callable metadata
    """
    # Determine the specific type of callable
    if inspect.isfunction(source):
        callable_type = "function"
    elif inspect.ismethod(source):
        callable_type = "instance method"
    elif inspect.isbuiltin(source):
        callable_type = "built-in"
    elif inspect.ismethoddescriptor(source):
        callable_type = "method descriptor"
    else:
        callable_type = "other callable"

    callable_info = {
        "type": callable_type,
        "name": getattr(source, '__name__', 'Unnamed'),
        "doc": inspect.getdoc(source) or "No documentation available.",
        "module": getattr(source, '__module__', None),
        "string_repr": repr(source),
        "signature": None,
        "source": None
    }

    # Attempt to get the signature
    try:
        signature = inspect.signature(source)
        callable_info["signature"] = str(signature)
    except (TypeError, ValueError):
        callable_info["signature"] = "Not available"

    # Attempt to get the source code if it's a regular function
    if callable_type == "function":
        try:
            callable_info["source"] = inspect.getsource(source)
        except Exception:
            callable_info["source"] = "Not available"

    return json.dumps(callable_info, indent=4)


def handle_module(source: Any) -> str:
    """Generate detailed JSON representation of a module.

    Args:
        source: Module object to inspect

    Returns:
        JSON string with module metadata
    """
    module_info = {
        "type": "module",
        "name": source.__name__,
        "doc": inspect.getdoc(source) or "No documentation available.",
        "functions": {},
        "classes": {},
        "variables": {},
        "submodules": {},
        "string_repr": repr(source)
    }

    for name, member in inspect.getmembers(source):
        if inspect.isfunction(member) and not name.startswith('_'):
            module_info["functions"][name] = {
                "doc": inspect.getdoc(member) or "No documentation available.",
                "signature": str(inspect.signature(member))
            }
        elif inspect.isclass(member) and not name.startswith('_'):
            module_info["classes"][name] = {
                "doc": inspect.getdoc(member) or "No documentation available."
            }
        elif inspect.ismodule(member) and not name.startswith('_'):
            module_info["submodules"][name] = {
                "doc": inspect.getdoc(member) or "No documentation available."
            }
        elif not (inspect.isroutine(member) or inspect.isclass(member) or inspect.ismodule(member)) and not name.startswith('_'):
            try:
                value = json.dumps(member)
            except:
                value = repr(member)

            module_info["variables"][name] = {
                "value": value,
                "type": type(member).__name__
            }

    return json.dumps(module_info, indent=4)


def handle_object(source: Any) -> str:
    """Generate detailed JSON representation of a generic object instance.

    Args:
        source: Object instance to inspect

    Returns:
        JSON string with instance metadata
    """
    # Try to serialize the object
    serialized = repr(source)
    if hasattr(source, 'dumps') and callable(getattr(source, 'dumps')):
        try:
            serialized = source.dumps()
        except:
            pass
    else:
        try:
            serialized = json.dumps(source)
        except:
            pass

    instance_info = {
        "type": "class_instance",
        "class": source.__class__.__name__,
        "doc": inspect.getdoc(source) or "No documentation available.",
        "attributes": {},
        "methods": {},
        "properties": {},
        "callable": callable(source),
        "call_signature": None,
        "string_repr": serialized
    }

    if instance_info["callable"]:
        try:
            instance_info["call_signature"] = str(inspect.signature(source.__call__))
        except (TypeError, ValueError):
            instance_info["call_signature"] = "Not available"

    # Use inspect.getmembers to safely retrieve object members
    members = inspect.getmembers(source, lambda a: not inspect.isroutine(a))
    methods = inspect.getmembers(source, inspect.isroutine)
    properties = [m for m in members if isinstance(m[1], property)]

    # Filter attributes, methods, and properties
    for name, member in members:
        if not name.startswith('__'):
            try:
                value = json.dumps(member)
            except:
                value = repr(member)

            instance_info["attributes"][name] = {
                "value": value,
                "type": type(member).__name__,
                "string_repr": repr(member)
            }

    for name, method in methods:
        if not name.startswith('__'):
            try:
                signature = str(inspect.signature(method))
            except (TypeError, ValueError):
                signature = "Not available"

            instance_info["methods"][name] = {
                "doc": inspect.getdoc(method) or "No documentation available.",
                "signature": signature
            }

    for name, prop in properties:
        instance_info["properties"][name] = {
            "doc": inspect.getdoc(prop) or "No documentation available."
        }

    return json.dumps(instance_info, indent=4)


# ============================================================================
# Main Entry Point
# ============================================================================

def get_text(source: Any) -> str:
    """Extract text from various sources.

    This is the main entry point that dispatches to appropriate handlers
    based on the type of source.

    Supported sources:
    - Python objects (classes, functions, modules, instances)
    - URLs (web pages, documents)
    - File paths (PDF, DOCX, ODT, HTML, text files)
    - Directory paths (generates tree structure)
    - Primitive types (lists, dicts, etc.)
    - Plain strings (returned as-is)

    Args:
        source: The source to extract text from

    Returns:
        Extracted text, cleaned of excessive newlines
    """
    # Handle non-string objects (Python objects)
    if not isinstance(source, str):
        # Classes
        if inspect.isclass(source):
            text = handle_class(source)

        # Callables (functions, methods, builtins)
        elif inspect.isfunction(source) or inspect.ismethod(source) or \
             inspect.isbuiltin(source) or inspect.ismethoddescriptor(source):
            text = handle_function(source)

        # Modules
        elif inspect.ismodule(source):
            text = handle_module(source)

        # Built-in types
        elif isinstance(source, (list, dict, tuple, set, int, float, bool)):
            try:
                text = json.dumps(source, indent=4)
            except:
                text = repr(source)

        # Generic objects
        elif isinstance(source, object):
            text = handle_object(source)

        # Fallback
        else:
            text = repr(source)

    # Handle strings (URLs, file paths, or plain text)
    elif source.startswith('http://') or source.startswith('https://'):
        text = handle_url(source)

    elif os.path.isfile(source):
        text = handle_file(source)

    elif os.path.isdir(source):
        text = handle_directory(source)

    else:
        # Plain string
        text = source

    return strip_newlines(text)
